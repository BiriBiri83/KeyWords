# -*- coding: utf-8 -*-

import pymorphy2
#import dawg_python
#import nltk
from many_stop_words import get_stop_words as stw
import openpyxl as op
liness=[]
parsewords=[]
wordslist=[]
line=''
i=0
punct=['.',',','«','»','"','!','?','—',':','(',')'] # знаки пунктуации
line1=''
print('Укажите путь')
path=input()
with open(path, 'r') as text:
    lines=text.readlines()
for a in lines:
    if a!='\n':
        line=line+a[:-1]+' ' # концы строк мне не нужны
#print(line)
line1=line
for i in range(len(line)): # цикл, который убирает ссылки из википедии в формате [х]
    if line[i]=='[':
        deletestr=''
        while line[i]!=']':
            deletestr=deletestr+line[i]
            i=i+1
        deletestr=deletestr+']'
        line1=line1.replace(deletestr,'')
for p in punct:
    line1=line1.replace(p,'')
line1=line1.replace(' '+' ',' ')    # убираю двойные пробелы, если есть
print(line1)
blah=pymorphy2.MorphAnalyzer()
texxt=line1.split(' ')
stwlist=list(stw('ru'))
for word in texxt:
    parsewords.append(blah.parse(word)[0].normal_form)
for pp in parsewords:
    flag=0
    if len(wordslist)!=0:
        for ww in wordslist:
            if pp==ww['word']:
                ww['count']=ww['count']+1
                flag=1        
    if flag==0 and pp not in stwlist:
        worddict=dict()
        worddict['word']=pp
        worddict['count']=1
        wordslist.append(worddict)  
        
wordslist.sort(key=lambda x: x['count'], reverse=True)         
print(wordslist)
wb = op.Workbook() # создаю книгу
ws = wb.active    
for i in range(len(wordslist)):
    if wordslist[i]['count']>=2: # слова, встречающиеся больше 1 раза
        ws[f'A{i+1}']=wordslist[i]['word']
        ws[f'B{i+1}']=wordslist[i]['count']
wb.save('Exportkeywords.xlsx') # сохраняю книгу